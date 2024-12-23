import openpyxl
from openpyxl.styles import PatternFill
import matplotlib


def rgb2hex(rgb):
    r, g, b, _a = rgb
    result = (r << 16) + (g << 8) + b
    return hex(result)

def get_color(before_value, after_value):
    if before_value == None or after_value == None:
        # white
        return "ffffff"
    cmap = matplotlib.colormaps['bwr']
    diff = before_value - after_value
    if diff > 0:
        diff = 128 - int((diff/before_value)*255/3)*6
    elif diff < 0:
        diff = 127 + int((-1*diff/before_value)*255/3)*6
    else:
        return "ffffff"
    color = rgb2hex(cmap(diff, bytes=True))[2:].zfill(6)
    return color

TABLE = './test.xlsx'
NEW_TABLE = "./new_test.xlsx"

# Keys to be enumerated
colMap = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O']
colMap = dict(zip(range(1, len(colMap)+1), colMap))

# Open file
wb = openpyxl.load_workbook(TABLE)
sheet = wb.active

# Fill the colors
for ridx, row in enumerate(sheet.iter_rows(min_row=2, min_col=2, values_only=True)):
    for cidx in range(0, len(row), 2):
        color = get_color(row[cidx], row[cidx+1])
        cor1 = colMap[cidx+1]+str(ridx+2)
        sheet[cor1].fill = PatternFill(start_color=color, fill_type='solid')
        cor2 = colMap[cidx+2]+str(ridx+2)
        sheet[cor2].fill = PatternFill(start_color=color, fill_type='solid')

wb.save(NEW_TABLE)
        
        



